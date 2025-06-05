import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
from utils.filling import create_driver, login, generate_urls, collect_data, get_tanggal_input


def save_to_excel_with_number_format(data, filename):
    """Menyimpan data ke Excel dengan format number yang benar"""
    if not data:
        print("❌ Tidak ada data untuk disimpan")
        return
    
    # Buat workbook baru
    wb = Workbook()
    ws = wb.active
    
    # Tambahkan data ke worksheet
    for row in data:
        ws.append(row)
    
    # Format kolom yang berisi angka (kolom 2 dan seterusnya biasanya berisi angka)
    # Asumsikan kolom pertama adalah nama bagian (text), sisanya adalah angka
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(2, ws.max_column + 1):  # Mulai dari kolom kedua
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                # Coba konversi ke float jika memungkinkan
                if cell.value and str(cell.value).replace('.', '').replace(',', '').replace('-', '').isdigit():
                    # Hapus koma jika ada (format angka Indonesia)
                    numeric_value = float(str(cell.value).replace(',', ''))
                    cell.value = numeric_value
                    cell.number_format = numbers.FORMAT_NUMBER
            except (ValueError, TypeError):
                # Jika tidak bisa dikonversi, biarkan sebagai text
                pass
    
    # Simpan file
    wb.save(filename)
    print(f"✅ Data berhasil disimpan dengan format number yang benar: {filename}")


def main():
    print("🚀 Memulai scraping data LOSS...")

    input_tanggal = get_tanggal_input()
    print(f"📅 Tanggal yang dipilih: {input_tanggal}")

    loss_urls = generate_urls("laporan/loss_bagian_cetak", input_tanggal)
    print(f"🔗 Generated {len(loss_urls)} URLs")

    driver = create_driver()

    try:
        login(driver)
        loss_data = collect_data(driver, loss_urls, "LOSS")

        print(f"\n📊 Data yang terkumpul: {len(loss_data)} baris")

        if not loss_data:
            print("❌ Tidak ada data yang berhasil dikumpulkan!")
            print("🔍 Kemungkinan penyebab:")
            print("  - Selector CSS tidak cocok dengan struktur HTML")
            print("  - Tanggal tidak memiliki data")
            print("  - Halaman web berubah struktur")
            print("  - Koneksi timeout")
            return

        os.makedirs("data", exist_ok=True)

        filename = f"data/filling-tem-{input_tanggal}.xlsx"

        if len(loss_data) > 0:
            print(f"📋 Struktur data pertama: {loss_data[0]}")
            print(f"📏 Jumlah kolom: {len(loss_data[0]) if loss_data else 0}")

            # Gunakan fungsi baru untuk menyimpan dengan format number
            save_to_excel_with_number_format(loss_data, filename)

            if os.path.exists(filename):
                file_size = os.path.getsize(filename)
                print(f"✅ Data LOSS berhasil disimpan di '{filename}'")
                print(f"📁 Ukuran file: {file_size} bytes")

                # Verifikasi dengan pandas
                df_test = pd.read_excel(filename, header=None)
                print(f"✅ Verifikasi: File berisi {len(df_test)} baris")
            else:
                print(f"❌ File tidak berhasil dibuat: {filename}")
        else:
            print("❌ Data kosong, file tidak dibuat")

    except Exception as e:
        print(f"❌ Terjadi kesalahan: {e}")
        import traceback
        traceback.print_exc()

    finally:
        print("\n🟢 Browser tetap terbuka untuk pemeriksaan manual.")
        print("💡 Tips debugging:")
        print("  1. Cek apakah halaman web berisi tabel data")
        print("  2. Inspect element untuk melihat struktur HTML tabel")
        print("  3. Pastikan tanggal memiliki data")
        print("  4. Cek console browser untuk error JavaScript")


if __name__ == "__main__":
    main()