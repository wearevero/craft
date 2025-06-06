import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
from utils.loss import create_driver, login, generate_urls, collect_data, get_tanggal_input
import re


def save_to_excel_with_number_format(data, filename):
    """Menyimpan data ke Excel dengan format number yang benar sesuai format asli dari web"""
    if not data:
        print("❌ Tidak ada data untuk disimpan")
        return
    
    # Buat workbook baru
    wb = Workbook()
    ws = wb.active
    
    # Tambahkan data ke worksheet
    for row in data:
        ws.append(row)
    
    # Format kolom yang berisi angka dengan mempertahankan format asli
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(2, ws.max_column + 1):  # Mulai dari kolom kedua
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if cell.value and str(cell.value).strip():
                cell_value_str = str(cell.value).strip()
                
                # Cek apakah ini adalah angka (termasuk desimal)
                if is_numeric_value(cell_value_str):
                    # Bersihkan nilai dari koma ribuan (jika ada format Indonesia)
                    cleaned_value = cell_value_str.replace(',', '')
                    
                    try:
                        # Konversi ke float untuk memastikan bisa diproses Excel
                        numeric_value = float(cleaned_value)
                        cell.value = numeric_value
                        
                        # Tentukan format berdasarkan nilai asli dari web
                        if '.' in cell_value_str:
                            # Hitung jumlah digit desimal dari nilai asli
                            decimal_places = len(cell_value_str.split('.')[-1])
                            
                            # Set format sesuai jumlah desimal yang ada
                            if decimal_places == 1:
                                cell.number_format = '0.0'
                            elif decimal_places == 2:
                                cell.number_format = '0.00'
                            elif decimal_places == 3:
                                cell.number_format = '0.000'
                            else:
                                # Untuk desimal lebih dari 3 digit
                                cell.number_format = '0.' + '0' * decimal_places
                        else:
                            # Jika tidak ada desimal, tampilkan sebagai integer
                            if numeric_value == int(numeric_value):
                                cell.number_format = '0'
                            else:
                                # Jika ada desimal tapi tidak terlihat di string asli
                                cell.number_format = '0.00'
                                
                    except (ValueError, TypeError):
                        # Jika tidak bisa dikonversi, biarkan sebagai text
                        pass
    
    # Simpan file
    wb.save(filename)
    print(f"✅ Data berhasil disimpan dengan format number yang benar: {filename}")


def is_numeric_value(value_str):
    """Mengecek apakah string adalah nilai numerik"""
    cleaned = value_str.replace(',', '')
    
    pattern = r'^-?\d+\.?\d*$'
    
    return bool(re.match(pattern, cleaned))


def main():
    print("🚀 Memulai scraping data LOSS...")

    input_tanggal = get_tanggal_input()
    print(f"📅 Tanggal yang dipilih: {input_tanggal}")

    loss_urls = generate_urls("laporan/loss_bagian_cetak", input_tanggal)
    print(f"🔗 Generated {len(loss_urls)} URLs")

    driver = create_driver()

    try:
        login(driver)
        loss_data = collect_data(driver, loss_urls, "LOSS")

        print(f"\n📊 Data yang terkumpul: {len(loss_data)} baris")

        if not loss_data:
            print("❌ Tidak ada data yang berhasil dikumpulkan!")
            print("🔍 Kemungkinan penyebab:")
            print("  - Selector CSS tidak cocok dengan struktur HTML")
            print("  - Tanggal tidak memiliki data")
            print("  - Halaman web berubah struktur")
            print("  - Koneksi timeout")
            return

        os.makedirs("data", exist_ok=True)

        filename = f"data/loss-{input_tanggal}.xlsx"

        if len(loss_data) > 0:
            print(f"📋 Struktur data pertama: {loss_data[0]}")
            print(f"📏 Jumlah kolom: {len(loss_data[0]) if loss_data else 0}")

            save_to_excel_with_number_format(loss_data, filename)

            if os.path.exists(filename):
                file_size = os.path.getsize(filename)
                print(f"✅ Data LOSS berhasil disimpan di '{filename}'")
                print(f"📁 Ukuran file: {file_size} bytes")

                df_test = pd.read_excel(filename, header=None)
                print(f"✅ Verifikasi: File berisi {len(df_test)} baris")
                
                print("\n🔍 Debug info - Sample data dari Excel:")
                for i, row in enumerate(df_test.head(3).iterrows()):
                    print(f"Baris {i+1}: {list(row[1])}")
                    
            else:
                print(f"❌ File tidak berhasil dibuat: {filename}")
        else:
            print("❌ Data kosong, file tidak dibuat")

    except Exception as e:
        print(f"❌ Terjadi kesalahan: {e}")
        import traceback
        traceback.print_exc()

    finally:
        print("\n🟢 Browser tetap terbuka untuk pemeriksaan manual.")
        print("💡 Tips debugging:")
        print("  1. Cek apakah halaman web berisi tabel data")
        print("  2. Inspect element untuk melihat struktur HTML tabel")
        print("  3. Pastikan tanggal memiliki data")
        print("  4. Cek console browser untuk error JavaScript")
        print("  5. Periksa format angka di Excel apakah sesuai dengan web")


if __name__ == "__main__":
    main()