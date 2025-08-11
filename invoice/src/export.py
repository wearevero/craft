# Export module for processed invoice data
import pandas as pd
import os
from pathlib import Path
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

class InvoiceExporter:
    def __init__(self, template_folder="template", output_folder="output"):
        self.template_folder = template_folder
        self.output_folder = output_folder
        self.setup_logging()
        self.ensure_directories()
        
    def setup_logging(self):
        """Setup logging untuk export process"""
        log_folder = "logs"
        os.makedirs(log_folder, exist_ok=True)
        log_file = os.path.join(log_folder, f"export_log_{datetime.now().strftime('%Y%m%d')}.log")
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def ensure_directories(self):
        """Membuat direktori yang diperlukan"""
        os.makedirs(self.template_folder, exist_ok=True)
        os.makedirs(self.output_folder, exist_ok=True)
    
    def find_processed_files(self):
        """Mencari file yang sudah diproses di folder template"""
        processed_files = []
        
        # Cari file processed_*.xlsx
        for file_path in Path(self.template_folder).glob("processed_*.xlsx"):
            processed_files.append(str(file_path))
        
        self.logger.info(f"Ditemukan {len(processed_files)} file yang sudah diproses")
        return processed_files
    
    def apply_excel_formatting(self, workbook, worksheet):
        """Menerapkan formatting pada Excel"""
        try:
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Border style
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # Apply header formatting
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Apply borders to all data cells
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                         min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            self.logger.info("Excel formatting applied successfully")
            
        except Exception as e:
            self.logger.error(f"Error applying formatting: {e}")
    
    def create_summary_sheet(self, workbook, all_data):
        """Membuat sheet summary dari semua data"""
        try:
            summary_ws = workbook.create_sheet(title="Summary")
            
            # Calculate summary statistics
            total_items = len(all_data)
            unique_pos = all_data['PO#'].nunique() if 'PO#' in all_data.columns else 0
            total_quantity = all_data["Q'ty"].sum() if "Q'ty" in all_data.columns else 0
            total_weight = all_data["Total w't"].sum() if "Total w't" in all_data.columns else 0
            
            # Create summary data
            summary_data = [
                ["RINGKASAN INVOICE", ""],
                ["", ""],
                ["Total Item", total_items],
                ["Unique PO Numbers", unique_pos],
                ["Total Quantity", total_quantity],
                ["Total Weight", total_weight],
                ["", ""],
                ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                summary_ws.cell(row=row_idx, column=1, value=label)
                summary_ws.cell(row=row_idx, column=2, value=value)
            
            # Format summary sheet
            title_font = Font(bold=True, size=14)
            summary_ws['A1'].font = title_font
            
            # Auto-adjust column widths
            for column in summary_ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = max_length + 2
                summary_ws.column_dimensions[column_letter].width = adjusted_width
            
            self.logger.info("Summary sheet created successfully")
            
        except Exception as e:
            self.logger.error(f"Error creating summary sheet: {e}")
    
    def export_single_file(self, processed_file, output_format="xlsx"):
        """Export single processed file dengan formatting"""
        try:
            self.logger.info(f"Exporting file: {processed_file}")
            
            # Read processed data
            df = pd.read_excel(processed_file)
            
            if df.empty:
                self.logger.warning(f"File {processed_file} kosong")
                return None
            
            # Create output filename
            base_name = Path(processed_file).stem
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_name = f"final_{base_name}_{timestamp}.{output_format}"
            output_path = os.path.join(self.output_folder, output_name)
            
            if output_format.lower() == "xlsx":
                # Export to Excel with formatting
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Invoice Data', index=False)
                    
                    # Get workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['Invoice Data']
                    
                    # Apply formatting
                    self.apply_excel_formatting(workbook, worksheet)
                    
                    # Create summary sheet
                    self.create_summary_sheet(workbook, df)
            
            else:
                # Export to CSV
                output_path = output_path.replace('.xlsx', '.csv')
                df.to_csv(output_path, index=False, encoding='utf-8-sig')
            
            self.logger.info(f"File exported to: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error exporting {processed_file}: {e}")
            return None
    
    def export_combined_file(self, processed_files, output_format="xlsx"):
        """Menggabungkan semua file dan export sebagai satu file"""
        try:
            self.logger.info("Creating combined export file")
            
            all_data = []
            file_info = []
            
            for file_path in processed_files:
                try:
                    df = pd.read_excel(file_path)
                    if not df.empty:
                        # Add source file column
                        df['Source_File'] = Path(file_path).stem
                        all_data.append(df)
                        file_info.append({
                            'file': Path(file_path).name,
                            'rows': len(df),
                            'processed_date': datetime.fromtimestamp(
                                os.path.getmtime(file_path)
                            ).strftime('%Y-%m-%d %H:%M:%S')
                        })
                except Exception as e:
                    self.logger.error(f"Error reading {file_path}: {e}")
            
            if not all_data:
                self.logger.warning("Tidak ada data untuk digabungkan")
                return None
            
            # Combine all data
            combined_df = pd.concat(all_data, ignore_index=True)
            
            # Create output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_name = f"combined_invoice_data_{timestamp}.{output_format}"
            output_path = os.path.join(self.output_folder, output_name)
            
            if output_format.lower() == "xlsx":
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    # Main data sheet
                    combined_df.to_excel(writer, sheet_name='Combined Data', index=False)
                    
                    # File info sheet
                    file_info_df = pd.DataFrame(file_info)
                    file_info_df.to_excel(writer, sheet_name='File Info', index=False)
                    
                    # Apply formatting to both sheets
                    workbook = writer.book
                    for sheet_name in writer.sheets:
                        worksheet = writer.sheets[sheet_name]
                        self.apply_excel_formatting(workbook, worksheet)
                    
                    # Create summary sheet
                    self.create_summary_sheet(workbook, combined_df)
            
            else:
                # Export to CSV
                output_path = output_path.replace('.xlsx', '.csv')
                combined_df.to_csv(output_path, index=False, encoding='utf-8-sig')
            
            self.logger.info(f"Combined file exported to: {output_path}")
            self.logger.info(f"Total records: {len(combined_df)}")
            
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error creating combined export: {e}")
            return None
    
    def export_data(self, export_type="combined", output_format="xlsx"):
        """Main export function"""
        self.logger.info("=== MEMULAI PROSES EXPORT ===")
        
        # Find processed files
        processed_files = self.find_processed_files()
        
        if not processed_files:
            self.logger.warning("Tidak ada file yang sudah diproses untuk diexport")
            return None
        
        exported_files = []
        
        if export_type == "individual":
            # Export each file individually
            for file_path in processed_files:
                result = self.export_single_file(file_path, output_format)
                if result:
                    exported_files.append(result)
        
        elif export_type == "combined":
            # Export as single combined file
            result = self.export_combined_file(processed_files, output_format)
            if result:
                exported_files.append(result)
        
        else:  # both
            # Export individually
            for file_path in processed_files:
                result = self.export_single_file(file_path, output_format)
                if result:
                    exported_files.append(result)
            
            # Export combined
            result = self.export_combined_file(processed_files, output_format)
            if result:
                exported_files.append(result)
        
        # Summary
        self.logger.info("=== RINGKASAN EXPORT ===")
        self.logger.info(f"File yang diproses: {len(processed_files)}")
        self.logger.info(f"File berhasil diexport: {len(exported_files)}")
        
        if exported_files:
            self.logger.info("File yang diexport:")
            for file in exported_files:
                self.logger.info(f"  - {file}")
        
        return exported_files
    
    def create_template_file(self):
        """Membuat file template kosong"""
        template_path = os.path.join(self.template_folder, "template.xlsx")
        
        # Template columns
        columns = ['PO#', 'Item', 'No.', 'Metal', "Q'ty", "Total w't", 'maklon', 'total']
        
        # Create empty dataframe with template columns
        template_df = pd.DataFrame(columns=columns)
        
        try:
            with pd.ExcelWriter(template_path, engine='openpyxl') as writer:
                template_df.to_excel(writer, sheet_name='Template', index=False)
                
                # Format template
                workbook = writer.book
                worksheet = writer.sheets['Template']
                self.apply_excel_formatting(workbook, worksheet)
            
            self.logger.info(f"Template file created: {template_path}")
            return template_path
            
        except Exception as e:
            self.logger.error(f"Error creating template: {e}")
            return None

def main():
    """Fungsi untuk testing export module"""
    exporter = InvoiceExporter()
    
    print("Testing Invoice Exporter...")
    
    # Create template if not exists
    template_path = exporter.create_template_file()
    if template_path:
        print(f"Template created: {template_path}")
    
    # Try to export data
    exported_files = exporter.export_data("combined", "xlsx")
    
    if exported_files:
        print(f"\nExport berhasil:")
        for file in exported_files:
            print(f"  - {file}")
    else:
        print("Tidak ada file untuk diexport")

if __name__ == "__main__":
    main()